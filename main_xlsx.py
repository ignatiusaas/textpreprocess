from openpyxl import load_workbook
import os
import glob
import tokenPDF as tkn
import string

import time
start_time = time.time()

blacklist = ['mail', 'ugm', 'ac', 'id']

#Get answers path
path = os.getcwd()+'\\answer\*.xlsx'

#Main code
for filepath in glob.glob(path):

    #Getting file name
    file_name = os.path.basename(filepath)
    file_name = os.path.splitext(file_name)[0]

    #Making new directory
    output_path = os.getcwd()+'\\'+str(file_name)
    try:
        os.mkdir(output_path)
    except FileExistsError:
        print("Output folder for " + str(file_name) + ' already exist, skipping...')

    #Loading all sheets
    wb = load_workbook(filepath)
    for sheet_title in wb.sheetnames:
        ws = wb[sheet_title]
        all_columns = list(ws.columns)

    #Fetching email column
    email_column = 1
    for column in range(1,ws.max_column):
        if 'email' in str(ws.cell(1,column).value).lower():
            break
        email_column += 1

    #Fetching response column
    response_column = 1
    for column in range(1,ws.max_column):
        if 'response' in str(ws.cell(1,column).value).lower():
            break
        response_column += 1

    #Extracting answers to txt
    ws_row = 2
    fn_i = 1
    for row in range(1,ws.max_row):
        if(ws.cell(row,1).value is None):
            break
        new_answ = ''

        #Fetching email for filename
        fn = ws.cell(ws_row,email_column).value
        fn = fn.translate(str.maketrans(string.punctuation, ' '*len(string.punctuation)))
        fn = ''.join(fn)
        fn = fn.split()
        fn = [i for i in fn if not i in blacklist]
        fn = ''.join(fn)
        #print(new_fn)

        #Fetching answers
        for column in range(response_column,ws.max_column):
            if(ws.cell(ws_row,column).value is None):
                break
            answ = ws.cell(ws_row,column).value
            new_answ = new_answ + answ
            #print(new_answ)
        ws_row += 1

        #Tokenize and lemmatize answers
        fin_ans = tkn.tokenPDF(new_answ,'')
        fin_ans = tkn.ligma(fin_ans)

        #Export output
        ouput = open(output_path + '\\' + str(fn_i).zfill(3) + '_' + str(fn)+'.txt', "w")
        ouput.write(' '.join(fin_ans))
        ouput.close
        fn_i += 1

#Print execution time
print("--- %s seconds ---" % (time.time() - start_time))
